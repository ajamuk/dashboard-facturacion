import os
import re
from collections import defaultdict
from pathlib import Path

import openpyxl
import requests
from google.oauth2 import service_account
from google.auth.transport.requests import Request


SHEETS = [
    "CUADRO MANDOS",
    "DETALLE FACTURACIÓN PARLA",
    "DETALLE FACTURACIÓN LAS ROSAS",
    "DETALLE FACTURACIÓN GETAFE",
    "TARIFAS PARLA",
    "TARIFAS LAS ROSAS",
]

MONTHS = [
    ("ene", "Enero"),
    ("feb", "Febrero"),
    ("mar", "Marzo"),
    ("abr", "Abril"),
    ("may", "Mayo"),
    ("jun", "Junio"),
    ("jul", "Julio"),
    ("ago", "Agosto"),
    ("sep", "Septiembre"),
    ("oct", "Octubre"),
    ("nov", "Noviembre"),
    ("dic", "Diciembre"),
]

MONTH_COLS = list(range(2, 14))
CENTERS = {"PARLA", "LAS ROSAS", "GETAFE", "TOTAL"}


def clean_text(value):
    if value is None:
        return ""
    return str(value).strip()


def number(value):
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("€", "").replace("%", "").replace(" ", "")
    text = text.replace(".", "").replace(",", ".") if "," in text else text
    if "??" in text or text in {"#DIV/0!", "#VALUE!", "??"}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def year_value(value):
    n = number(value)
    if n and 1900 <= n <= 2100:
        return int(n)
    return None


def sheet_title_to_center(sheet_name):
    if "PARLA" in sheet_name:
        return "PARLA"
    if "LAS ROSAS" in sheet_name:
        return "LAS ROSAS"
    if "GETAFE" in sheet_name:
        return "GETAFE"
    return ""


def load_workbook_rows_from_xlsx(path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    result = {}
    for sheet_name in SHEETS:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows():
            rows.append([cell.value for cell in row])
        result[sheet_name] = rows
    return result


def load_workbook_rows_from_google():
    sheet_id = os.environ["GOOGLE_SHEET_ID"]
    credentials_file = os.environ["GOOGLE_SERVICE_ACCOUNT_FILE"]
    scopes = ["https://www.googleapis.com/auth/spreadsheets.readonly"]
    creds = service_account.Credentials.from_service_account_file(credentials_file, scopes=scopes)
    creds.refresh(Request())
    headers = {"Authorization": f"Bearer {creds.token}"}
    result = {}
    for sheet_name in SHEETS:
        range_name = requests.utils.quote(f"'{sheet_name}'", safe="")
        url = (
            f"https://sheets.googleapis.com/v4/spreadsheets/{sheet_id}/values/"
            f"{range_name}?valueRenderOption=UNFORMATTED_VALUE"
        )
        response = requests.get(url, headers=headers, timeout=30)
        response.raise_for_status()
        result[sheet_name] = response.json().get("values", [])
    return result


def load_source_rows():
    if os.environ.get("GOOGLE_SHEET_ID") and os.environ.get("GOOGLE_SERVICE_ACCOUNT_FILE"):
        return load_workbook_rows_from_google()
    xlsx = os.environ.get("SOURCE_XLSX")
    if not xlsx:
        raise RuntimeError("Configura SOURCE_XLSX o GOOGLE_SHEET_ID + GOOGLE_SERVICE_ACCOUNT_FILE")
    return load_workbook_rows_from_xlsx(Path(xlsx))


def find_rows(rows, title):
    title_norm = title.upper()
    matches = []
    for idx, row in enumerate(rows):
        if any(clean_text(cell).upper() == title_norm for cell in row):
            matches.append(idx)
    return matches


def parse_regular_metric(rows, title, metric):
    starts = find_rows(rows, title)
    if not starts:
        return []
    start = starts[0]
    next_starts = [i for i, row in enumerate(rows[start + 1 :], start + 1) if any(clean_text(c).upper() in {
        "CLIENTES ACTIVOS", "GASTO MEDIO POR SOCIO", "ALTAS", "BAJAS",
        "OCUPACIÓN CLASES", "LIFETIME VALUE", "PERMANENCIA MEDIA"
    } for c in row)]
    end = next_starts[0] if next_starts else len(rows)
    records = []
    current_year = None
    for row in rows[start:end]:
        if len(row) < 14:
            continue
        y = year_value(row[0])
        if y:
            current_year = y
        center = clean_text(row[1]).upper()
        if not current_year or center not in CENTERS:
            continue
        for col, (_, month_name) in zip(MONTH_COLS, MONTHS):
            val = number(row[col] if col < len(row) else None)
            if val is None:
                continue
            records.append(
                {
                    "metric": metric,
                    "year": current_year,
                    "month": month_name,
                    "month_index": MONTH_COLS.index(col) + 1,
                    "center": center,
                    "value": val,
                }
            )
    return records


def parse_bajas(rows):
    starts = find_rows(rows, "BAJAS")
    if not starts:
        return []
    start = starts[0]
    end_candidates = find_rows(rows, "OCUPACIÓN CLASES")
    end = end_candidates[0] if end_candidates else len(rows)
    count_cols = [2, 4, 6, 8, 10, 12, 14, 16, 18, 20, 22, 24]
    records = []
    current_year = None
    for row in rows[start:end]:
        y = year_value(row[0] if len(row) > 0 else None)
        if y:
            current_year = y
        center = clean_text(row[1] if len(row) > 1 else "").upper()
        if not current_year or center not in CENTERS:
            continue
        for i, col in enumerate(count_cols):
            count = number(row[col] if col < len(row) else None)
            pct = number(row[col + 1] if col + 1 < len(row) else None)
            if count is None:
                continue
            records.append(
                {
                    "metric": "bajas",
                    "year": current_year,
                    "month": MONTHS[i][1],
                    "month_index": i + 1,
                    "center": center,
                    "value": count,
                    "percentage": pct,
                }
            )
    return records


def parse_cuadro_mandos(rows):
    records = []
    for title, metric in [
        ("FACTURACIÓN", "facturacion"),
        ("CLIENTES ACTIVOS", "clientes_activos"),
        ("GASTO MEDIO POR SOCIO", "gasto_medio"),
        ("ALTAS", "altas"),
        ("OCUPACIÓN CLASES", "ocupacion_clases"),
        ("LIFETIME VALUE", "ltv"),
        ("PERMANENCIA MEDIA", "permanencia"),
    ]:
        records.extend(parse_regular_metric(rows, title, metric))
    records.extend(parse_bajas(rows))
    return records


def parse_service_details(sheet_name, rows):
    center = sheet_title_to_center(sheet_name)
    records = []
    service = None
    for row in rows:
        first = clean_text(row[0] if row else "")
        y = year_value(row[0] if row else None)
        if first and not y and first.upper() not in {"TOTAL"}:
            if not any(first.lower().startswith(m[1].lower()) for m in MONTHS):
                service = first
        if not y or not service:
            continue
        for col, (_, month_name) in zip(MONTH_COLS, MONTHS):
            val = number(row[col] if col < len(row) else None)
            if val is None:
                continue
            records.append(
                {
                    "center": center,
                    "service": service,
                    "year": y,
                    "month": month_name,
                    "month_index": MONTH_COLS.index(col) + 1,
                    "value": val,
                }
            )
    return records


def parse_tariffs(sheet_name, rows):
    center = sheet_title_to_center(sheet_name)
    records = []
    for year_col, name_col in [(2, 1), (7, 6)]:
        title = clean_text(rows[0][year_col] if rows and len(rows[0]) > year_col else "")
        match = re.search(r"(20\d{2})", title)
        if not match:
            continue
        year = int(match.group(1))
        price_col = year_col
        classes_col = year_col + 1
        for row in rows[3:9]:
            name = clean_text(row[name_col] if len(row) > name_col else "")
            if not name:
                continue
            price = number(row[price_col] if len(row) > price_col else None)
            classes = number(row[classes_col] if len(row) > classes_col else None)
            if price is None or classes is None:
                continue
            records.append(
                {
                    "center": center,
                    "year": year,
                    "tariff": name,
                    "price": price,
                    "classes": classes,
                    "price_per_class": price / classes if classes else None,
                }
            )
    return records


def latest_year_month(records, metric="facturacion"):
    filtered = [r for r in records if r["metric"] == metric and r["center"] != "TOTAL"]
    if not filtered:
        return None, None
    latest = max(filtered, key=lambda r: (r["year"], r["month_index"]))
    return latest["year"], latest["month_index"]


def summarize(metric_records, service_records):
    year, month_index = latest_year_month(metric_records)
    centers = sorted({r["center"] for r in metric_records if r["center"] != "TOTAL"})
    annual = defaultdict(float)
    latest_month = defaultdict(float)
    metrics_latest = defaultdict(dict)
    for r in metric_records:
        if r["center"] == "TOTAL":
            continue
        if r["metric"] == "facturacion":
            annual[(r["year"], r["center"])] += r["value"]
            if r["year"] == year and r["month_index"] == month_index:
                latest_month[r["center"]] += r["value"]
        if r["year"] == year and r["month_index"] == month_index:
            metrics_latest[r["center"]][r["metric"]] = r["value"]
    services_latest = defaultdict(float)
    for r in service_records:
        if r["year"] == year:
            services_latest[(r["center"], r["service"])] += r["value"]
    return {
        "latest": {"year": year, "month_index": month_index, "month": MONTHS[month_index - 1][1] if month_index else None},
        "centers": centers,
        "annual_revenue": [
            {"year": y, "center": c, "value": v}
            for (y, c), v in sorted(annual.items())
        ],
        "latest_month_revenue": [
            {"center": c, "value": latest_month.get(c, 0)}
            for c in centers
        ],
        "latest_metrics": [
            {"center": center, **values}
            for center, values in sorted(metrics_latest.items())
        ],
        "services_year": [
            {"center": c, "service": s, "value": v}
            for (c, s), v in sorted(services_latest.items())
        ],
    }


def load_dashboard_payload():
    rows_by_sheet = load_source_rows()
    metric_records = parse_cuadro_mandos(rows_by_sheet["CUADRO MANDOS"])
    service_records = []
    for sheet_name in [
        "DETALLE FACTURACIÓN PARLA",
        "DETALLE FACTURACIÓN LAS ROSAS",
        "DETALLE FACTURACIÓN GETAFE",
    ]:
        service_records.extend(parse_service_details(sheet_name, rows_by_sheet[sheet_name]))
    tariffs = []
    for sheet_name in ["TARIFAS PARLA", "TARIFAS LAS ROSAS"]:
        tariffs.extend(parse_tariffs(sheet_name, rows_by_sheet[sheet_name]))
    return {
        "summary": summarize(metric_records, service_records),
        "metrics": metric_records,
        "services": service_records,
        "tariffs": tariffs,
        "quality_notes": [
            "La app lee la hoja original en modo solo lectura.",
            "Los valores no numericos o con marcas tipo ?? se omiten en las series.",
            "Los calculos visuales del dashboard se hacen sobre datos normalizados en cache.",
        ],
    }
